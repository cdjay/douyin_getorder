"""
Excel数据导入模块
负责读取、解析和导入Excel文件到数据库
"""
import os
import logging
from datetime import datetime, date
import openpyxl
from typing import List, Dict, Any

logger = logging.getLogger(__name__)


class ExcelImporter:
    """Excel数据导入器"""
    
    def __init__(self, db_manager):
        """
        初始化导入器
        
        Args:
            db_manager: 数据库管理器
        """
        self.db_manager = db_manager
    
    def read_excel_file(self, file_path: str) -> List[Dict[str, Any]]:
        """
        读取Excel文件（单sheet）
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            Excel数据列表（每行一个字典）
        """
        try:
            # 使用data_only=True模式，更稳定
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # 调试信息
            logger.info(f"  打开工作表: {ws.title}")
            logger.info(f"  总行数: {ws.max_row}")
            
            # 获取表头（第一行）
            headers = [cell.value for cell in ws[1]]
            logger.info(f"  表头: {headers[:5]}...")  # 只打印前5个
            
            # 读取数据行
            data = []
            empty_rows = 0
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # 检查空行
                if not any(row):
                    empty_rows += 1
                    continue
                
                row_dict = {}
                for j, value in enumerate(row):
                    if j < len(headers):
                        row_dict[headers[j]] = value
                data.append(row_dict)
                
                # 每1000行打印一次进度
                if len(data) % 1000 == 0:
                    logger.info(f"  已读取 {len(data)} 行...")
            
            logger.info(f"  读取完成: {len(data)} 行有效数据 (跳过 {empty_rows} 行空数据)")
            wb.close()
            return data
            
        except Exception as e:
            logger.error(f"读取Excel失败 {file_path}: {e}", exc_info=True)
            raise
    
    def read_sheet(self, sheet) -> List[Dict[str, Any]]:
        """
        读取单个工作表
        
        Args:
            sheet: openpyxl工作表对象
            
        Returns:
            数据列表
        """
        try:
            # 调试信息
            logger.info(f"  打开工作表: {sheet.title}")
            logger.info(f"  总行数: {sheet.max_row}")
            
            # 获取表头（第一行）
            headers = [cell.value for cell in sheet[1]]
            logger.info(f"  表头: {headers[:5]}...")  # 只打印前5个
            
            # 读取数据行
            data = []
            empty_rows = 0
            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                # 检查空行
                if not any(row):
                    empty_rows += 1
                    continue
                
                row_dict = {}
                for j, value in enumerate(row):
                    if j < len(headers):
                        row_dict[headers[j]] = value
                data.append(row_dict)
                
                # 每1000行打印一次进度
                if len(data) % 1000 == 0:
                    logger.info(f"  已读取 {len(data)} 行...")
            
            logger.info(f"  读取完成: {len(data)} 行有效数据 (跳过 {empty_rows} 行空数据)")
            return data
        except Exception as e:
            logger.error(f"读取工作表失败: {e}", exc_info=True)
            return []
    
    def parse_sales_data(self, excel_data: List[Dict[str, Any]], file_name: str) -> List[Dict[str, Any]]:
        """
        解析售卖明细Excel数据
        
        Args:
            excel_data: Excel原始数据
            file_name: 文件名
            
        Returns:
            数据库格式的数据列表
        """
        parsed_data = []
        
        for row in excel_data:
            # 构建数据库记录
            record = {
                'order_id': str(row.get('所属订单ID', '')) if row.get('所属订单ID') else None,
                'sub_order_id': str(row.get('子订单ID', '')) if row.get('子订单ID') else None,
                'coupon_status': str(row.get('券码状态', '')) if row.get('券码状态') else None,
                'verification_time': self._parse_datetime(row.get('核销时间')),
                'actual_receipt': self._parse_float(row.get('订单实收')),
                'sale_amount': self._parse_float(row.get('售卖金额')),
                'merchant_subsidy': self._parse_float(row.get('商家货款出资补贴')),
                'product_payment': self._parse_float(row.get('商品实付')),
                'platform_subsidy': self._parse_float(row.get('平台补贴')),
                'platform_discount_detail': str(row.get('平台补贴优惠明细', '')) if row.get('平台补贴优惠明细') else None,
                'software_fee': self._parse_float(row.get('软件服务费')),
                'talent_commission': self._parse_float(row.get('达人佣金')),
                'increment_commission': self._parse_float(row.get('增量宝佣金')),
                'preset_price': self._parse_float(row.get('预售价(只针对酒旅商家)')),
                'booking_surcharge': self._parse_float(row.get('预约加价(只针对酒旅商家)')),
                'software_fee_rate': str(row.get('软件服务费率', '')) if row.get('软件服务费率') else None,
                'sales_role': str(row.get('带货角色', '')) if row.get('带货角色') else None,
                'deal_channel': str(row.get('成交渠道', '')) if row.get('成交渠道') else None,
                'owner_nickname': str(row.get('订单归属人昵称(字段如果获取不到就默认展示商家，具体金额以账单为主)', '商家')) if row.get('订单归属人昵称(字段如果获取不到就默认展示商家，具体金额以账单为主)') else '商家',
                'owner_uid': str(row.get('订单归属人uid', '')) if row.get('订单归属人uid') else None,
                'raw_excel': row,  # 原始数据
                'file_name': file_name,
                'import_time': datetime.now()
            }
            
            parsed_data.append(record)
        
        logger.info(f"解析完成: {len(parsed_data)} 条记录")
        return parsed_data
    
    def parse_travel_booking_data(self, excel_data: List[Dict[str, Any]], sheet_name: str) -> List[Dict[str, Any]]:
        """
        解析旅行社预约明细数据
        
        Args:
            excel_data: Excel原始数据
            sheet_name: 工作表名
            
        Returns:
            数据库格式的数据列表
        """
        parsed_data = []
        
        # 打印前3条数据的原始值，用于调试
        for i, row in enumerate(excel_data[:3]):
            logger.info(f"  示例数据 {i+1}: 订单编号={row.get('订单编号')}, 出行日期={row.get('出行日期')}, 类型={type(row.get('出行日期'))}")
        
        for row in excel_data:
            # 构建数据库记录
            record = {
                'order_number': str(row.get('订单编号', '')) if row.get('订单编号') else None,
                'travel_date': self._parse_date(row.get('出行日期')),
                'booking_status': self._parse_booking_status(sheet_name),
                'booking_count': self._parse_int(row.get('预约份数', 1)),
                'raw_excel': row,  # 原始数据
                'import_time': datetime.now()
            }
            
            parsed_data.append(record)
        
        logger.info(f"解析完成: {len(parsed_data)} 条记录 (sheet: {sheet_name})")
        return parsed_data
    
    def _parse_datetime(self, value) -> datetime:
        """解析日期时间"""
        if value is None:
            return None
        if isinstance(value, datetime):
            return value
        # Excel日期是数字，需要转换
        try:
            return datetime.fromordinal(int(value) + 693594)
        except:
            return None
    
    def _parse_date(self, value) -> date:
        """解析日期（支持多种格式）"""
        if value is None:
            return None
        
        # 已经是datetime或date对象
        if isinstance(value, datetime):
            logger.debug(f"  日期格式: datetime -> {value}")
            return value.date()
        if isinstance(value, date):
            logger.debug(f"  日期格式: date -> {value}")
            return value
        
        # 字符串格式：尝试解析常见日期格式
        if isinstance(value, str):
            date_str = value.strip()
            logger.debug(f"  日期格式: string -> {date_str}")
            
            if not date_str:
                logger.warning(f"  日期为空字符串")
                return None
            
            # 格式1: YYYY-MM-DD
            if '-' in date_str:
                try:
                    result = datetime.strptime(date_str, '%Y-%m-%d').date()
                    logger.debug(f"  解析成功 (YYYY-MM-DD): {result}")
                    return result
                except:
                    try:
                        result = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S').date()
                        logger.debug(f"  解析成功 (YYYY-MM-DD HH:MM:SS): {result}")
                        return result
                    except:
                        pass
            
            # 格式2: YYYY/MM/DD
            if '/' in date_str:
                try:
                    result = datetime.strptime(date_str, '%Y/%m/%d').date()
                    logger.debug(f"  解析成功 (YYYY/MM/DD): {result}")
                    return result
                except:
                    try:
                        result = datetime.strptime(date_str, '%Y/%m/%d %H:%M:%S').date()
                        logger.debug(f"  解析成功 (YYYY/MM/DD HH:MM:SS): {result}")
                        return result
                    except:
                        pass
        
        # Excel数字日期（Excel序列日期）
        try:
            result = datetime.fromordinal(int(value) + 693594).date()
            logger.debug(f"  日期格式: Excel数字 {value} -> {result}")
            return result
        except:
            logger.warning(f"  无法解析日期: {value} (type: {type(value)})")
            return None
    
    def _parse_int(self, value, default: int = 0) -> int:
        """解析整数"""
        if value is None:
            return default
        try:
            return int(value)
        except:
            return default
    
    def _parse_float(self, value) -> float:
        """解析浮点数"""
        if value is None:
            return 0.0
        try:
            return float(value)
        except:
            return 0.0
    
    def _parse_booking_status(self, sheet_name: str) -> str:
        """根据工作表名解析预约状态"""
        if '已预约' in sheet_name:
            return '已预约'
        elif '已完成' in sheet_name:
            return '已完成'
        else:
            return sheet_name  # 默认使用sheet名
    
    def import_sales_excel(self, file_path: str, file_name: str) -> int:
        """
        导入售卖明细Excel
        
        Args:
            file_path: Excel文件路径
            file_name: 文件名
            
        Returns:
            int: 导入的记录数
        """
        logger.info(f"开始导入售卖明细: {file_name}")
        
        try:
            # 读取Excel文件
            excel_data = self.read_excel_file(file_path)
            
            # 解析数据
            parsed_data = self.parse_sales_data(excel_data, file_name)
            
            # 保存到数据库
            saved_count = self.db_manager.save_excel_orders(parsed_data, file_name)
            
            logger.info(f"✅ 售卖明细导入完成: {saved_count} 条记录")
            return saved_count
            
        except Exception as e:
            logger.error(f"导入售卖明细失败 {file_name}: {e}")
            raise
    
    def import_travel_booking_excel(self, file_path: str, file_name: str) -> int:
        """
        导入旅行社预约明细Excel
        
        Args:
            file_path: Excel文件路径
            file_name: 文件名（仅用于日志）
            
        Returns:
            int: 导入的记录数
        """
        logger.info(f"开始导入旅行社预约明细: {file_name}")
        
        total_imported = 0
        
        try:
            # 使用data_only=True模式，更稳定
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            logger.info(f"  发现 {len(wb.worksheets)} 个工作表")
            
            # 遍历所有工作表
            for sheet in wb.worksheets:
                sheet_name = sheet.title
                logger.info(f"  检查工作表: {sheet_name}")
                
                # 跳过说明类sheet
                if '说明' in sheet_name:
                    logger.info(f"  跳过非数据工作表: {sheet_name}")
                    continue
                
                # 读取数据
                excel_data = self.read_sheet(sheet)
                
                # 解析数据
                parsed_data = self.parse_travel_booking_data(excel_data, sheet_name)
                
                # 保存到数据库
                saved_count = self.db_manager.save_travel_bookings(parsed_data)
                total_imported += saved_count
            
            wb.close()
            logger.info(f"✅ 旅行社预约明细导入完成: {total_imported} 条记录")
            return total_imported
            
        except Exception as e:
            logger.error(f"导入旅行社预约明细失败 {file_name}: {e}", exc_info=True)
            raise
    
    def scan_and_import(self, data_dir: str = 'data') -> int:
        """
        扫描并导入所有Excel文件（支持两种类型）
        
        导入成功后删除Excel文件
        
        Args:
            data_dir: data文件夹路径
            
        Returns:
            int: 导入的记录数
        """
        total_imported = 0
        
        if not os.path.exists(data_dir):
            logger.warning(f"目录不存在: {data_dir}")
            return 0
        
        # 扫描所有.xlsx文件
        for filename in os.listdir(data_dir):
            if not filename.endswith('.xlsx'):
                continue
            
            file_path = os.path.join(data_dir, filename)
            
            try:
                # 判断Excel类型
                if filename.startswith('售卖明细_'):
                    # 类型1：售卖明细
                    count = self.import_sales_excel(file_path, filename)
                elif filename.startswith('旅行社预约明细_'):
                    # 类型2：旅行社预约明细
                    count = self.import_travel_booking_excel(file_path, filename)
                else:
                    logger.warning(f"未知的Excel类型: {filename}")
                    continue
                
                total_imported += count
                
                if count > 0:
                    # 导入成功，删除文件
                    os.remove(file_path)
                    logger.info(f"✅ 已删除已导入文件: {filename}")
            except Exception as e:
                logger.error(f"导入文件失败 {filename}: {e}")
        
        return total_imported
